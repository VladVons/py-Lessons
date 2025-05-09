# Created: 2025.03.11
# Author: Vladimir Vons <VladVons@gmail.com>
#
# Tester: https://messente.com/sms-length-calculator

import os
import asyncio
from openpyxl import load_workbook
#
from Inc.Misc.SmsGW import TSmsGW
from Inc.DbList import TDbList
from Inc.Log import TLog, TEchoConsole, TEchoFile
from Inc.Var.Dict import DeepGet


LogFile = 'SmsGW.log'
Log = TLog([TEchoConsole(), TEchoFile(LogFile)])

FmtText = '''
{Name},
В Україні створено великий каталог вживаних комп’ютерів - FindWares.
https://it.findwares.com/uk/?adv=1-{Id}#a-content
'''

# FmtText = '''
# {Name},
# В Україні створено великий каталог вживаних комп’ютерів.
# Весь товар продавців зібрано в одну базу.
# Обирай кращі пропозиції!
# https://it.findwares.com/uk/?xid=1-{Id}#a-content
# '''

# FmtText = '''
# Name: {Name},
# Id: {Id}
# https://it.findwares.com/uk
# '''

class TAdver():
    def __init__(self):
        self.Sleep = 120
        Connect = {
            'aUrl': 'http://192.168.2.208:8080',
            'aUser': 'sms',
            'aPassw': 'F63VRxdn'
        }
        self.SmsGW = TSmsGW(**Connect)

    def LogPhones(self) -> set:
        Res = set()
        if (os.path.exists(LogFile)):
            with open(LogFile, 'r', encoding='utf8') as F:
                for xLine in F.readlines():
                    Parts = [xPart.strip() for xPart in xLine.split(',')]
                    _Day, _Hour, _No, _Level, _Type, Phone, Status, _SmsId = Parts
                    if (Phone) and (Status in ('200', '202')):
                        Res.add(Phone)
        return Res

    async def SendDict(self, aData: dict):
        Text = FmtText.format(Id=aData['id'], Name=aData['name']).strip()
        Res = await self.SmsGW.Send(aData['phones'], Text)
        Log.Print(1, 'i', f"{aData['phones'][0]}, {Res['status']}, {DeepGet(Res, 'data.id')}")
        await asyncio.sleep(self.Sleep)

    async def SendDbl(self, aName: str):
        Sent = self.LogPhones()

        Dbl = TDbList()
        Dbl.Load(aName)
        for Rec in Dbl:
            Phone = Rec.phone.replace('+38', '')
            if (Phone not in Sent) and (not Phone.startswith('-')):
                Data = {
                    'id': Rec.id,
                    'phones': [Phone],
                    'name': Rec.GetField('name', 'Привіт'),
                }
                await self.SendDict(Data)

    async def SendXlsx(self, aName: str):
        def GetField(aName: str):
            nonlocal headers, row
            return row[headers[aName]]

        Sent = self.LogPhones()

        wb = load_workbook(aName)
        ws = wb.active

        headers = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
        headers = {name.lower(): idx for idx, name in enumerate(headers)}

        for row in ws.iter_rows(min_row=2, values_only=True):
            Phone = GetField("телефон 1")
            if (Phone not in Sent):
                Obl = GetField('область')
                if (Obl == 'Хмельницька область'):
                    Data = {
                        'id': GetField('id автора'),
                        'phones': [Phone],
                        'name':  GetField("ім'я"),
                    }
                    await self.SendDict(Data)


    async def Test(self):
        Text = FmtText.format(Name='Володимир', Id=12345)
        #Text = 'Hello'
        Phones = ['0976646510']
        Res = await self.SmsGW.Send(Phones, Text)
        #Res = await self.SmsGW.Status('KeYuIsLPEq7sLl-5ro1BF')
        #Res = await self.SmsGW.Logs()
        #Res = await self.SmsGW.Health()
        print(f'Status: {Res["status"]}, Data: {Res["data"]}')

Adver = TAdver()
Task = Adver.SendXlsx('OLX-RepairPC.xlsx')
#Task = Adver.SendDbl('test.dbl.json')
#Task = Adver.Test()
asyncio.run(Task)

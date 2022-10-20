'''
Pyrhon example
Excel

vladvons@gmail.com
2022.10.29
'''

from datetime import datetime
#
from openpyxl import Workbook
#from openpyxl.utils import get_column_letter
#from openpyxl.styles import Font
#from openpyxl.comments import Comment
#
from Data import Ruska_24


def Task_01(aData: dict):
    for x in aData:
        print(x)

def Task_02(aData: dict):
    for x in aData:
        print('Name: %s' % (x['Name']))

def Task_03(aData: dict):
    for x in aData:
        DOB = x['DOB']
        DateDOB = datetime.strptime(DOB, '%Y-%m-%d').date()
        Now = datetime.now()
        print('Name: %s, Age: %s' % (x['Name'], Now.year - DateDOB.year))

def Task_10(aData: dict):
    FileName = 'Task_10.xlsx'
    SheetName = 'Vons family'

    WB = Workbook()
    WS = WB.create_sheet(title = SheetName)
    for PersonNo, Person in enumerate(aData):
        for FieldNo, Field in enumerate(Person):
            WS.cell(PersonNo + 1, FieldNo + 1).value = Person[Field]

    #WS.freeze_panes = WS.cell(2, 1)
    WB.save(FileName)


#Task_01(Ruska_24)
#Task_02(Ruska_24)
#Task_03(Ruska_24)
Task_10(Ruska_24)
